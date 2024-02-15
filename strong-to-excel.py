import csv
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import logging
import sys
import argparse
import time

log = logging.getLogger(__name__)
parser = argparse.ArgumentParser(prog='strong-to-excel', description='program to import data from Strong app to Excel sheet')
parser.add_argument('-l', '--logging', choices=['DEBUG', 'INFO', 'WARNING', "ERROR", "CRITICAL"], default='DEBUG', help="do no print to std output")
parser.add_argument('-o', '--output', default=f'strong-{time.strftime("%Y%m%d")}.xlsx', help="output file name, default - strong-<timestamp>.xlsx")
parser.add_argument('-i', '--input', default=f'strong.csv', help="input file name, default - strong.csv")
args = parser.parse_args()

def setup_logger():
    numeric_level = getattr(logging, args.logging.upper(), None)
    log.setLevel(numeric_level)
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(formatter)
    log.addHandler(handler)


def getYearMonthString(date : datetime.date) -> str:
    monthName = {
        1: "January",
        2: "February",
        3: "March",
        4: "April",
        5: "May",
        6: "June",
        7: "July",
        8: "August",
        9: "September",
        10: "October",
        11: "November",
        12: "December"
    }
    return monthName[date.month] + " " + str(date.year)

def getDate(stringDate):
   return datetime.strptime(stringDate, "%Y-%m-%d %H:%M:%S").date()

def minimizeRow(row):
    return {
        "workoutName": row["Workout Name"],
        "exerciseName": row["Exercise Name"],
        "setOrder": row["Set Order"],
        "weight": 0 if row["Weight"] == '' else float(row["Weight"]),
        "reps": 0 if row["Reps"] == '' else int(row["Reps"]),
    }

def transformTraining(training):
    groupedExercises = {}

    for item in training[1]:
        if item['exerciseName'] not in groupedExercises:
            groupedExercises[item['exerciseName']] = []
        groupedExercises[item['exerciseName']].append({'setOrder': item['setOrder'], 'reps': item['reps'], 'weight': item['weight']})

    return {
        'name': training[1][0]['workoutName'],
        'date': training[0],
        'exercises': groupedExercises
    }

def getGroupedTrainings():
    groupedTrainings = {}
    log.debug(f"opening file: {args.input}")
    with open(args.input, newline="") as csvfile:
        file = csv.DictReader(csvfile, delimiter=";")
        for row in file:
            date = row["Date"]
            if date not in groupedTrainings:
                groupedTrainings[date] = []
            groupedTrainings[date].append(minimizeRow(row))
    for training in groupedTrainings.items():
        groupedTrainings[training[0]] = transformTraining(training)
    return groupedTrainings

def adjustWidth(sheet):
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 0.92
        sheet.column_dimensions[column].width = adjusted_width




def main():
    setup_logger()
    workbook = Workbook()
    sheet = workbook.active

    yellowFills = [PatternFill(start_color='facb5c',
                    end_color='facb5c',
                    fill_type='solid'),
                PatternFill(start_color='f5d68e',
                    end_color='f5d68e',
                    fill_type='solid')
                ]
    blueFills = [PatternFill(start_color='4a8af7',
                    end_color='4a8af7',
                    fill_type='solid'),
                PatternFill(start_color='86b0f7',
                    end_color='86b0f7',
                    fill_type='solid')
                ]

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    Fills = [yellowFills, blueFills]
    fillsCounter = 0
    groupedTrainings = getGroupedTrainings()
    columnCounter = 1
    rowCounter = 1
    maxRow = 1
    currentRow = 1
    monthMergeRow = 1

    oldYearMonthString = ""
    for training in groupedTrainings.items():
        currentFills = Fills[fillsCounter%2]
        fillsCounter += 1
        date = getDate(training[1]['date'])
        yearMonthString = getYearMonthString(date)
        if (yearMonthString != oldYearMonthString):
            sheet.merge_cells(start_row=monthMergeRow, end_row=monthMergeRow, start_column=1, end_column=columnCounter)
            sheet.cell(row=monthMergeRow, column = 1).font = Font(size="30")
            sheet.row_dimensions[monthMergeRow].height = 35
            oldYearMonthString = yearMonthString 
            columnCounter = 1
            currentRow = maxRow+2
            rowCounter = currentRow

        sheet.cell(row=rowCounter, column=columnCounter).value = yearMonthString
        monthMergeRow = rowCounter
        rowCounter += 1
        sheet.cell(row=rowCounter, column=columnCounter).value = training[1]['name']
        sheet.cell(row=rowCounter, column=columnCounter).font = Font(bold=True, size=15)
        sheet.cell(row=rowCounter, column=columnCounter).fill = currentFills[0]
        sheet.cell(row=rowCounter, column=columnCounter).border = thin_border
        sheet.cell(row=rowCounter, column=columnCounter+1).value = str(date)
        sheet.cell(row=rowCounter, column=columnCounter+1).font = Font(bold=True, size=15)
        sheet.cell(row=rowCounter, column=columnCounter+1).fill = currentFills[0]
        sheet.cell(row=rowCounter, column=columnCounter+1).border = thin_border
        sheet.merge_cells(start_row=rowCounter, end_row=rowCounter, start_column=columnCounter+1, end_column=columnCounter+2) 
        sheet.row_dimensions[rowCounter].height = 20
        rowCounter += 1
        sheet.cell(row=rowCounter, column=columnCounter).value = "exercise"
        sheet.cell(row=rowCounter, column=columnCounter).fill = currentFills[0]
        sheet.cell(row=rowCounter, column=columnCounter).border = thin_border
        sheet.cell(row=rowCounter, column=columnCounter+1).value = "reps"
        sheet.cell(row=rowCounter, column=columnCounter+1).fill = currentFills[0]
        sheet.cell(row=rowCounter, column=columnCounter+1).border = thin_border
        sheet.cell(row=rowCounter, column=columnCounter+2).value = "weight"
        sheet.cell(row=rowCounter, column=columnCounter+2).fill = currentFills[0]
        sheet.cell(row=rowCounter, column=columnCounter+2).border = thin_border


        exerciseFillCounter = 1
        for exercise in training[1]['exercises'].items():
            counter = 0
            
            for details in exercise[1]:
                rowCounter += 1
                if counter == 0:
                    sheet.cell(row=rowCounter, column=columnCounter).value = exercise[0]
                    sheet.cell(row=rowCounter, column=columnCounter).fill = currentFills[(exerciseFillCounter)%2]
                    sheet.cell(row=rowCounter, column=columnCounter).border = thin_border
                counter += 1
                sheet.cell(row=rowCounter, column=columnCounter+1).value = details['reps']
                sheet.cell(row=rowCounter, column=columnCounter+1).fill = currentFills[(exerciseFillCounter)%2]
                sheet.cell(row=rowCounter, column=columnCounter+1).border = thin_border
                sheet.cell(row=rowCounter, column=columnCounter+2).value = details['weight']
                sheet.cell(row=rowCounter, column=columnCounter+2).fill = currentFills[(exerciseFillCounter)%2]
                sheet.cell(row=rowCounter, column=columnCounter+2).border = thin_border
            exerciseFillCounter += 1
            sheet.merge_cells(start_row=rowCounter-counter+1, end_row=rowCounter, start_column=columnCounter, end_column=columnCounter) 
            sheet.cell(row=rowCounter-counter+1, column = columnCounter).alignment = Alignment(vertical='center', horizontal='center')
            sheet.cell(row=rowCounter-counter+1, column = columnCounter).fill = currentFills[(exerciseFillCounter-1)%2]
            sheet.cell(row=rowCounter-counter+1, column = columnCounter).border = thin_border
                
        columnCounter += 4
        
        if(rowCounter > maxRow):
            maxRow = rowCounter
        rowCounter = currentRow
        
    sheet.merge_cells(start_row=monthMergeRow, end_row=monthMergeRow, start_column=1, end_column=columnCounter)
    sheet.cell(row=monthMergeRow, column = 1).font = Font(size="30")
    sheet.row_dimensions[monthMergeRow].height = 35

    adjustWidth(sheet)
    log.debug(f"saving file: {args.output}")
    workbook.save(filename=args.output)
    

if __name__ == "__main__":
    main()