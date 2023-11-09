import csv
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Alignment


def getYearMonthString(date : datetime.date) -> str:
    return str(date.year) + "-" + str(date.month)

def getDate(stringDate):
   return datetime.strptime(stringDate, "%Y-%m-%d %H:%M:%S").date()

def minimizeRow(row):
    return {
        "workoutName": row["Workout Name"],
        "exerciseName": row["Exercise Name"],
        "setOrder": row["Set Order"],
        "weight": 0 if row["Weight"] == '' else float(row["Weight"]),
        "reps": 0 if row["Reps"] == '' else int(row["Reps"]),
    }

def transformTraining(training):
    groupedExercises = {}

    for item in training[1]:
        if item['exerciseName'] not in groupedExercises:
            groupedExercises[item['exerciseName']] = []
        groupedExercises[item['exerciseName']].append({'setOrder': item['setOrder'], 'reps': item['reps'], 'weight': item['weight']})

    return {
        'name': training[1][0]['workoutName'],
        'date': training[0],
        'exercises': groupedExercises
    }

def getGroupedTrainings():
    groupedTrainings = {}
    with open("strong.csv", newline="") as csvfile:
        file = csv.DictReader(csvfile, delimiter=";")
        for row in file:
            date = row["Date"]
            if date not in groupedTrainings:
                groupedTrainings[date] = []
            groupedTrainings[date].append(minimizeRow(row))
    for training in groupedTrainings.items():
        groupedTrainings[training[0]] = transformTraining(training)
    return groupedTrainings


workbook = Workbook()
sheet = workbook.active

groupedTrainings = getGroupedTrainings()
columnCounter = 1
rowCounter = 1
maxRow = 1
currentRow = 1
monthMergeRow = 1
# flag = True
oldYearMonthString = ""
for training in groupedTrainings.items():
    date = getDate(training[1]['date'])
    yearMonthString = getYearMonthString(date)
    # if(yearMonthString == "2022-11" or yearMonthString == "2022-12" or yearMonthString == "2023-1"):
    if(yearMonthString != oldYearMonthString):
        sheet.merge_cells(start_row=monthMergeRow, end_row=monthMergeRow, start_column=1, end_column=columnCounter)
        sheet.cell(row=monthMergeRow, column = 1).alignment = Alignment(vertical='center', horizontal='center')
        oldYearMonthString = yearMonthString 
        columnCounter = 1
        currentRow = maxRow+2
        rowCounter = currentRow
        
        # rowCounter = maxRow #TODO tu jest cos nie tak

    sheet.cell(row=rowCounter, column=columnCounter).value = yearMonthString
    monthMergeRow = rowCounter
    rowCounter += 1
    sheet.cell(row=rowCounter, column=columnCounter).value = training[1]['name']
    sheet.cell(row=rowCounter, column=columnCounter+1).value = str(date)
    rowCounter += 1
    sheet.cell(row=rowCounter, column=columnCounter).value = "exercise"
    sheet.cell(row=rowCounter, column=columnCounter+1).value = "reps"
    sheet.cell(row=rowCounter, column=columnCounter+2).value = "weight"

    for exercise in training[1]['exercises'].items():
        counter = 0
        for details in exercise[1]:
            rowCounter += 1
            if counter == 0:
                sheet.cell(row=rowCounter, column=columnCounter).value = exercise[0]
            counter += 1
            sheet.cell(row=rowCounter, column=columnCounter+1).value = details['reps']
            sheet.cell(row=rowCounter, column=columnCounter+2).value = details['weight']
        sheet.merge_cells(start_row=rowCounter-counter+1, end_row=rowCounter, start_column=columnCounter, end_column=columnCounter) 
        sheet.cell(row=rowCounter-counter+1, column = columnCounter).alignment = Alignment(vertical='center', horizontal='center')

    columnCounter += 4
    
    if(rowCounter > maxRow):
        maxRow = rowCounter
    rowCounter = currentRow
    
    
    
for col in sheet.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 0.92
    sheet.column_dimensions[column].width = adjusted_width

workbook.save(filename="test.xlsx")